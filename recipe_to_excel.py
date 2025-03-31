import requests
import json
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

def get_recipe(recipe_name, token):
    """
    调用智能体API获取菜谱信息
    
    参数:
        recipe_name: 要查询的菜谱名称
        token: 个人访问令牌
    
    返回:
        解析后的菜谱信息
    """
    # API请求URL
    url = "https://api.coze.cn/open_api/v2/chat"
    
    # 设置请求头
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    
    # 根据示例代码构造请求体
    payload = {
        "bot_id": "7487100580821893160",
        "user": f"user_{abs(hash(recipe_name)) % 10000000}",  # 用户标识
        "query": f"{recipe_name}",  # 修改查询内容，明确要求菜谱信息
        "stream": False
    }
    
    print(f"发送请求到: {url}")
    print(f"请求体: {json.dumps(payload, ensure_ascii=False, indent=2)}")
    
    try:
        # 发送请求
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()  # 检查请求是否成功
        
        print(f"响应状态码: {response.status_code}")
        
        # 解析响应数据
        response_data = response.json()
        
        # 为了调试，打印完整响应数据（可能会很长）
        print(f"响应数据: {json.dumps(response_data, ensure_ascii=False, indent=2)[:500]}...（内容较长已截断）")
        
        # 从响应中提取消息
        messages = response_data.get('messages', [])
        
        if not messages:
            print("响应中没有messages字段")
            return None
            
        # 打印所有消息类型，帮助诊断问题
        print("消息类型列表:")
        for i, message in enumerate(messages):
            role = message.get('role', '未知角色')
            msg_type = message.get('type', '未知类型')
            print(f"消息 {i+1}: 角色 = {role}, 类型 = {msg_type}")
        
        # 查找assistant角色且type为answer的消息内容
        for message in messages:
            if message.get('role') == 'assistant' and message.get('type') == 'answer':
                message_content = message.get('content', '')
                print(f"提取的answer内容: {message_content[:500]}...（内容较长已截断）")
                
                # 尝试解析JSON内容
                try:
                    recipe_data = json.loads(message_content)
                    return recipe_data
                except json.JSONDecodeError as e:
                    print(f"JSON解析错误: {e}")
                    print(f"原始内容: {message_content}")
                    # 如果不是JSON格式，直接返回文本内容
                    return message_content
        
        print("未找到answer类型的助手回复，找到的消息类型有:")
        for message in messages:
            print(f"角色: {message.get('role')}, 类型: {message.get('type')}")
        
        return "未找到answer类型的助手回复"
        
    except requests.exceptions.RequestException as e:
        print(f"请求失败: {e}")
        return f"API请求失败: {str(e)}"
    except Exception as e:
        print(f"处理响应时出错: {e}")
        import traceback
        traceback.print_exc()  # 打印详细的堆栈跟踪
        return f"处理响应时出错: {str(e)}"

def save_to_excel(recipe_data, excel_path, current_id):
    """
    将菜谱数据保存到Excel表格
    
    参数:
        recipe_data: 菜谱数据字典
        excel_path: Excel文件路径
        current_id: 当前菜谱的ID
    """
    try:
        # 获取当前时间
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 从recipe_data中获取数据
        name = recipe_data.get('name', '')
        cook_time = recipe_data.get('cook_time', '')
        calories = recipe_data.get('calories', '')
        image = recipe_data.get('image', '')
        description = recipe_data.get('description', '')
        # 将列表和字典类型的字段转换为JSON字符串
        steps = json.dumps(recipe_data.get('steps', []), ensure_ascii=False)
        tools = json.dumps(recipe_data.get('tools', []), ensure_ascii=False)
        prep_steps = json.dumps(recipe_data.get('prep_steps', []), ensure_ascii=False)
        tips = json.dumps(recipe_data.get('tips', []), ensure_ascii=False)
        difficulty = recipe_data.get('difficulty', '')
        
        # 创建新行数据
        new_row = {
            'id': current_id,
            'name': name,
            'cook_time': cook_time,
            'calories': calories,
            'image': image,
            'description': description,
            'steps': steps,
            'tools': tools,
            'prep_steps': prep_steps,
            'tips': tips,
            'difficulty': difficulty,
            'created_at': current_time,
            'updated_at': current_time
        }
        
        # 加载Excel文件（保留所有工作表）
        book = load_workbook(excel_path)
        
        # 检查是否存在recipes工作表
        if 'recipes' not in book.sheetnames:
            # 如果不存在，创建新的工作表
            sheet = book.create_sheet('recipes')
            # 添加表头 - 更新表头，添加新字段
            sheet.append(['id', 'name', 'cook_time', 'calories', 'image', 'description', 'steps', 'tools', 'prep_steps', 'tips', 'difficulty', 'created_at', 'updated_at'])
        else:
            # 如果存在，获取该工作表
            sheet = book['recipes']
            
            # 检查是否需要更新表头（添加新字段）
            header_row = [cell.value for cell in sheet[1]]
            required_headers = ['id', 'name', 'cook_time', 'calories', 'image', 'description', 'steps', 'tools', 'prep_steps', 'tips', 'difficulty', 'created_at', 'updated_at']
            
            # 检查是否缺少某些表头
            missing_headers = [header for header in required_headers if header not in header_row]
            if missing_headers:
                print(f"表头缺少以下字段：{', '.join(missing_headers)}")
                print("将在现有表头后添加缺失字段")
                
                # 添加缺失的列头
                for header in missing_headers:
                    sheet.cell(row=1, column=len(header_row) + 1).value = header
                    header_row.append(header)
        
        # 添加新行
        # 获取当前表头
        current_headers = [cell.value for cell in sheet[1]]
        
        # 准备新行数据，按表头顺序
        row_data = []
        for header in current_headers:
            if header in new_row:
                row_data.append(new_row[header])
            else:
                row_data.append('')  # 对于未知字段，添加空值
                
        # 添加新行
        sheet.append(row_data)
        
        # 保存Excel文件
        book.save(excel_path)
        print(f"成功将菜谱 '{name}' 保存到Excel")
        return True
    except Exception as e:
        print(f"保存到Excel时出错: {e}")
        return False

def get_last_id(excel_path):
    """
    获取Excel中最后一个ID
    
    参数:
        excel_path: Excel文件路径
    
    返回:
        最后一个ID，如果没有则返回0
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(excel_path):
            return 0
        
        # 加载Excel文件
        book = load_workbook(excel_path)
        
        # 检查是否存在recipes工作表
        if 'recipes' not in book.sheetnames:
            return 0
        
        # 获取工作表
        sheet = book['recipes']
        
        # 获取行数
        row_count = sheet.max_row
        
        # 如果只有表头或为空，返回0
        if row_count <= 1:
            return 0
        
        # 获取最后一行的ID
        last_id = sheet.cell(row=row_count, column=1).value
        
        # 如果最后一行的ID为空，从倒数第二行开始查找非空ID
        if last_id is None:
            for row in range(row_count-1, 0, -1):
                last_id = sheet.cell(row=row, column=1).value
                if last_id is not None:
                    break
        
        # 如果找不到有效ID，返回0
        return int(last_id) if last_id is not None else 0
    except Exception as e:
        print(f"获取最后ID时出错: {e}")
        return 0

def main():
    # 使用提供的令牌
    token = "pat_YTZ66uuJRvL42sxcb8HEfTRMSuSIkjuHcjCX3jDOwnyYroO49dlKRhQDPnUiRJ17"
    
    # 文件路径
    recipes_file = "家常菜菜谱名称.txt"
    excel_path = "data/database.xlsx"
    
    # 从文件中读取菜谱名称
    with open(recipes_file, 'r', encoding='utf-8') as file:
        recipe_names = [line.strip() for line in file.readlines() if line.strip()]
    
    # 获取最后一个ID
    last_id = get_last_id(excel_path)
    print(f"Excel中的最后ID: {last_id}")
    
    # 为每个菜谱调用API并保存到Excel
    for index, recipe_name in enumerate(recipe_names, 1):
        print(f"\n{'-' * 50}")
        print(f"处理第{index}个菜谱: {recipe_name}")
        print(f"{'-' * 50}")
        
        # 调用API获取菜谱数据
        recipe_data = get_recipe(recipe_name, token)
        
        print(f"获取到的菜谱数据类型: {type(recipe_data)}")
        if recipe_data is None:
            print(f"获取菜谱 '{recipe_name}' 失败: 返回了None")
            continue
            
        if isinstance(recipe_data, str):
            print(f"获取菜谱 '{recipe_name}' 失败: 返回了字符串而非JSON对象")
            print(f"返回内容: {recipe_data}")
            continue
            
        if isinstance(recipe_data, dict):
            # 检查返回的字典是否包含必要的字段
            required_fields = ['name', 'steps']
            missing_fields = [field for field in required_fields if field not in recipe_data]
            
            if missing_fields:
                print(f"获取菜谱 '{recipe_name}' 返回的数据缺少必要字段: {', '.join(missing_fields)}")
                print(f"返回的字段有: {', '.join(recipe_data.keys())}")
                continue
                
            # 保存到Excel
            current_id = last_id + index
            try:
                success = save_to_excel(recipe_data, excel_path, current_id)
                if success:
                    print(f"成功保存菜谱 '{recipe_name}' 到Excel")
                else:
                    print(f"保存菜谱 '{recipe_name}' 到Excel失败")
            except Exception as e:
                print(f"保存菜谱 '{recipe_name}' 到Excel时发生异常: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"获取菜谱 '{recipe_name}' 失败: 返回了意外的数据类型 {type(recipe_data)}")
    
    print("\n所有菜谱处理完成")

if __name__ == "__main__":
    main() 